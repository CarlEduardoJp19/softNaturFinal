import json
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from ..decorators import admin_required
from ..models import Pedido, Usuario
from productos.models import Calificacion

@admin_required
def gstUsuarios(request):
    usuarios = Usuario.objects.all().order_by('id')  # opcional: ordenados por id
    
    # Paginación
    paginator = Paginator(usuarios, 10)  # 10 usuarios por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, "usuarios/gstUsuarios.html", {
        'page_obj': page_obj,
        'usuarios': page_obj.object_list  # lista de usuarios de la página actual
    })

@admin_required
def cambiar_estado_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    usuario.is_active = not usuario.is_active
    usuario.save()
    return redirect('usuarios:gstUsuarios')  # Cambia si tu ruta tiene otro nombre

@admin_required
def agregar_usuario(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        correo = request.POST.get('email')
        telefono = request.POST.get('phone_number')
        rol = request.POST.get('rol')
        password = request.POST.get('password')  # Asegúrate de tener un input para la contraseña
        # Validar si el email ya existe
        if Usuario.objects.filter(email=correo).exists():
            messages.error(request, 'Ya existe un usuario con ese correo.')
            return render(request, 'usuarios/gstUsuarios.html')
        # Crear usuario usando create_user para manejar contraseña
        Usuario.objects.create_user(
            email=correo,
            nombre=nombre,
            phone_number=telefono,
            rol=rol,
            password=password,
            is_active=True
        )

        messages.success(request, 'Usuario agregado correctamente.')
        return redirect('usuarios:gstUsuarios')

    return render(request, 'usuarios/agregar.html')

def editar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)

    if request.method == "POST":
        usuario.nombre = request.POST.get("nombre")
        usuario.email = request.POST.get("email")
        usuario.phone_number = request.POST.get("phone_number")
        usuario.rol = request.POST.get("rol")
        usuario.save()

        messages.success(request, "Usuario actualizado correctamente.")
        return redirect("usuarios:gstUsuarios")  # 👈 ajusta al nombre de tu vista/listado

    messages.error(request, "Método no permitido")
    return redirect("usuarios:gstUsuarios")

@admin_required
def pedidos_view(request):
    # 🔹 Obtener todos los pedidos pagados (sin filtros excesivos)
    pedidos = (
        Pedido.objects
        .filter(pago=True, total__gt=0)  # ← Añade total > 0 también
        .order_by('-fecha_creacion')
    )

    # 🔹 Paginación
    paginator = Paginator(pedidos, 25)  # 25 pedidos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 🔹 Conteos por estado (solo pedidos pagados)
    pedidos_pendientes = Pedido.objects.filter(estado="pendiente", pago=True).count()
    pedidos_enviados = Pedido.objects.filter(estado="enviado", pago=True).count()
    pedidos_entregados = Pedido.objects.filter(estado="entregado", pago=True).count()

    # 🔹 Total pedidos pagados
    total_pedidos = pedidos.count()

    return render(request, "usuarios/gst_pedidos.html", {
        "page_obj": page_obj,
        "pedidos": page_obj.object_list,
        "pedidos_pendientes": pedidos_pendientes,
        "pedidos_enviados": pedidos_enviados,
        "pedidos_entregados": pedidos_entregados,
        "total_pedidos": total_pedidos,
    })

@admin_required
@require_http_methods(["POST"])
def cambiar_estado_pedido(request, pedido_id):
    try:
        # Obtener el pedido
        pedido = get_object_or_404(Pedido, id=pedido_id)
        
        # Obtener el nuevo estado del body
        data = json.loads(request.body)
        nuevo_estado = data.get('estado')
        
        # Validar que el estado sea válido
        estados_validos = ['pendiente', 'enviado', 'entregado']
        if nuevo_estado not in estados_validos:
            return JsonResponse({
                'success': False,
                'message': 'Estado no válido'
            }, status=400)
        
        # Actualizar el estado
        pedido.estado = nuevo_estado
        pedido.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Estado actualizado a {nuevo_estado}',
            'nuevo_estado': nuevo_estado
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error al procesar los datos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)
    
def detalle_pedido(request, pedido_id):
    try:
        pedido = get_object_or_404(Pedido, id=pedido_id)
        usuario = getattr(pedido, 'usuario', None)

        # Obtener los items del pedido
        items = pedido.items.all()
        productos = []

        for item in items:
            producto = item.producto

            # Detectar correctamente el campo del nombre
            nombre_producto = getattr(producto, 'nombProduc', None) or getattr(producto, 'nombre', 'Producto')

            productos.append({
                'nombre': nombre_producto,
                'cantidad': item.cantidad,
                'precio': float(getattr(producto, 'precio', 0)),
                'subtotal': float(item.cantidad * getattr(producto, 'precio', 0))
            })

        # Respuesta JSON
        return JsonResponse({
            'success': True,
            'pedido': {
                'id': pedido.id,
                'usuario': getattr(usuario, 'nombre', 'N/A') if usuario else 'N/A',
                'email': getattr(usuario, 'email', 'N/A') if usuario else 'N/A',
                'telefono': getattr(usuario, 'telefono', 'N/A') if usuario else 'N/A',
                'direccion': getattr(usuario, 'direccion', 'N/A') if usuario else 'N/A',
                'estado': getattr(pedido, 'estado', 'Pendiente'),
                'pago': pedido.pago,
                'pagado': pedido.pago,  # Por si acaso lo usas en otro lugar
                'fecha': pedido.fecha_creacion.strftime('%d/%m/%Y') if hasattr(pedido, 'fecha_creacion') else 'N/A',
                'hora': pedido.fecha_creacion.strftime('%H:%M:%S') if hasattr(pedido, 'fecha_creacion') else '',
                'metodo_pago': getattr(pedido, 'metodo_pago', 'N/A'),
                'total': float(getattr(pedido, 'total', 0)),
                'productos': productos
            }
        })

    except Exception as e:
        import traceback
        print(f"Error completo: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@admin_required
def informe_calificaciones(request):
    calificaciones = Calificacion.objects.select_related('usuario', 'producto').all()

    tipo = request.GET.get('tipo')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if tipo:
        calificaciones = calificaciones.filter(servicio__tipo=tipo)
    if desde:
        calificaciones = calificaciones.filter(fecha_creacion__gte=desde)
    if hasta:
        calificaciones = calificaciones.filter(fecha_creacion__lte=hasta)

    # Exportar a Excel
    if request.GET.get('exportar') == 'excel':
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "Calificaciones"

        # Encabezado principal de la tienda (NARANJA)
        ws.merge_cells('A1:E1')
        ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
        ws.row_dimensions[1].height = 30

        # Subtítulo (NARANJA OSCURO)
        ws.merge_cells('A2:E2')
        ws['A2'] = 'INFORME DE CALIFICACIONES'
        ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
        ws.row_dimensions[2].height = 25

        # Fecha de generación
        ws.merge_cells('A3:E3')
        ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A3'].font = Font(italic=True, size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 20

        # Fila vacía
        ws.append([])

        # Encabezados de columnas (fila 5)
        encabezados = [
            'Usuario',
            'Puntuación Servicio',
            'Puntuación Producto',
            'Comentario',
            'Fecha'
        ]
        ws.append(encabezados)

        # Estilo para encabezados de columnas (AMARILLO)
        header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
        header_font = Font(bold=True, color='000000')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, 6):
            cell = ws.cell(row=5, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos de calificaciones
        row_num = 6
        for c in calificaciones:
            ws.append([
                c.usuario.nombre if c.usuario else 'Anónimo',
                c.puntuacion_servicio,
                c.puntuacion_productos,
                c.comentario,
                c.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ])

            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border

            row_num += 1

        # Ajustar ancho de columnas
        column_widths = [25, 20, 20, 40, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="calificaciones.xlsx"'
        wb.save(response)
        return response

    return render(request, 'usuarios/informe_calificaciones.html', {
        'calificaciones': calificaciones
    })

def aprobar_comentario(request, id):
    calificacion = get_object_or_404(Calificacion, id=id)
    calificacion.aprobado = True
    calificacion.save()
    return redirect('usuarios:informe_calificaciones')

def rechazar_comentario(request,id):
    calificacion = get_object_or_404(Calificacion, id=id)
    calificacion.delete()
    return redirect('usuarios:informe_calificaciones')

def informe_ventas(request):
    # 🔹 Tomamos pedidos que ya tengan pago confirmado
    pedidos_list = Pedido.objects.filter(
        pago=True,  # ya fue pagado
        total__gt=0  # que tengan monto
    ).order_by('-fecha_creacion')

    # 🔹 Paginación (25 por página)
    paginator = Paginator(pedidos_list, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 🔹 Totales
    total_ventas = pedidos_list.aggregate(Sum('total'))['total__sum'] or 0
    total_pedidos = pedidos_list.count()

    # 🔹 Contexto
    context = {
        'page_obj': page_obj,
        'total_ventas': total_ventas,
        'total_pedidos': total_pedidos,
    }

    return render(request, 'usuarios/ventas.html', context)

