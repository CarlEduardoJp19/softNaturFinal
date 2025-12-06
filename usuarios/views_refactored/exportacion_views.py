from openpyxl import Workbook
from ..models import Usuario, Pedido, Devolucion, Producto
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
import openpyxl
from django.shortcuts import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from usuarios.decorators import admin_required

def exportar_usuarios_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    # Encabezado principal de la tienda (NARANJA)
    ws.merge_cells('A1:E1')
    ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    ws.row_dimensions[1].height = 30

    # Subtítulo (NARANJA OSCURO)
    ws.merge_cells('A2:E2')
    ws['A2'] = 'LISTADO DE USUARIOS'
    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    ws.row_dimensions[2].height = 25

    # Fecha de generación
    ws.merge_cells('A3:E3')
    ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Fila vacía
    ws.append([])

    # Encabezados de columnas (fila 5)
    encabezados = ["ID", "Nombre", "Email", "Rol", "Teléfono"]
    ws.append(encabezados)

    # Estilo para encabezados de columnas (AMARILLO)
    header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, 6):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Datos de usuarios
    row_num = 6
    for u in Usuario.objects.all():
        ws.append([u.id, u.nombre, u.email, u.rol, u.phone_number])
        
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
        
        row_num += 1

    # Ajustar ancho de columnas
    column_widths = [10, 30, 35, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
    wb.save(response)

    return response

def exportar_pedidos_excel(request):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Encabezado principal (NARANJA)
    ws.merge_cells('A1:F1')
    ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    ws.row_dimensions[1].height = 30

    # Subtítulo (NARANJA OSCURO)
    ws.merge_cells('A2:F2')
    ws['A2'] = 'LISTADO DE PEDIDOS'
    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    ws.row_dimensions[2].height = 25

    # Fecha de generación
    ws.merge_cells('A3:F3')
    ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Fila vacía
    ws.append([])

    # Encabezados de columnas (AMARILLO)
    encabezados = ['ID', 'Cliente', 'Email', 'Fecha', 'Estado', 'Total']
    ws.append(encabezados)

    header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, 7):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Filtros desde GET
    estado = request.GET.get('estado')
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    # Consulta base
    pedidos = Pedido.objects.filter(pago=True).order_by('-fecha_creacion')
    
    if estado:
        pedidos = pedidos.filter(estado=estado)
    if mes:
        pedidos = pedidos.filter(fecha_creacion__month=mes)
    if anio:
        pedidos = pedidos.filter(fecha_creacion__year=anio)

    # Colores por estado
    estado_colors = {
        'pendiente': 'FCF3CF',   # Amarillo claro
        'enviado': 'D6EAF8',     # Azul claro
        'entregado': 'D5F4E6',   # Verde claro
    }

    # Datos
    row_num = 6
    total = 0
    for pedido in pedidos:
        ws.append([
            pedido.id,
            pedido.usuario.nombre if pedido.usuario else 'Sin usuario',
            pedido.usuario.email if pedido.usuario else 'Sin email',
            pedido.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            pedido.estado.capitalize(),
            float(pedido.total)
        ])
        total += float(pedido.total)

        # Color según estado
        estado_lower = pedido.estado.lower()
        row_fill = PatternFill(
            start_color=estado_colors.get(estado_lower, 'FFFFFF'),
            end_color=estado_colors.get(estado_lower, 'FFFFFF'),
            fill_type='solid'
        )

        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.fill = row_fill
        row_num += 1

    # Fila de total
    ws.append(['', '', '', '', 'TOTAL:', total])
    total_fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    for col in range(1, 7):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = total_fill

    # Ajustar ancho de columnas
    column_widths = [10, 25, 30, 20, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
    wb.save(response)
    return response

def exportar_ventas_excel(request):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Encabezado principal (NARANJA)
    ws.merge_cells('A1:D1')
    ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    ws.row_dimensions[1].height = 30

    # Subtítulo (NARANJA OSCURO)
    ws.merge_cells('A2:D2')
    ws['A2'] = 'INFORME DE VENTAS'
    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    ws.row_dimensions[2].height = 25

    # Fecha de generación
    ws.merge_cells('A3:D3')
    ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Fila vacía
    ws.append([])

    # Encabezados de columnas (AMARILLO)
    encabezados = ['ID', 'Cliente', 'Fecha', 'Monto']
    ws.append(encabezados)

    header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, 5):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Filtros desde GET
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    # Consulta base
    pedidos = Pedido.objects.filter(pago=True, total__gt=0).order_by('-fecha_creacion')
    
    if mes:
        pedidos = pedidos.filter(fecha_creacion__month=mes)
    if anio:
        pedidos = pedidos.filter(fecha_creacion__year=anio)

    # Datos
    row_num = 6
    total = 0
    for pedido in pedidos:
        ws.append([
            pedido.id,
            pedido.usuario.nombre if pedido.usuario else 'Sin usuario',
            pedido.fecha_creacion.strftime('%d/%m/%Y'),
            float(pedido.total)
        ])
        total += float(pedido.total)

        for col in range(1, 5):
            ws.cell(row=row_num, column=col).border = thin_border
        row_num += 1

    # Fila de total (AMARILLO CLARO)
    ws.append(['', '', 'TOTAL:', total])
    total_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.fill = total_fill

    # Ajustar ancho de columnas
    column_widths = [10, 30, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="informe_ventas.xlsx"'
    wb.save(response)
    return response

@admin_required
def exportar_dashboard_excel(request):
    # Obtener los mismos filtros del dashboard
    tiempo = request.GET.get("tiempo")
    mes_especifico = request.GET.get("mes_especifico", "")
    anio_especifico = request.GET.get("anio_especifico", "")
    
    hoy = timezone.now()

    # Aplicar los mismos filtros que en el dashboard
    pedidos = Pedido.objects.filter(
        pago=True,
        fecha_creacion__year__gte=2025
    )

    devoluciones = Devolucion.objects.filter(
        fecha_solicitud__year__gte=2025
    )

    # FILTRO DE TIEMPO
    if tiempo == "semana":
        inicio_semana = hoy - timezone.timedelta(days=hoy.weekday())
        pedidos = pedidos.filter(fecha_creacion__gte=inicio_semana)
        devoluciones = devoluciones.filter(fecha_solicitud__gte=inicio_semana)

    elif tiempo == "mes":
        pedidos = pedidos.filter(
            fecha_creacion__year=hoy.year,
            fecha_creacion__month=hoy.month
        )
        devoluciones = devoluciones.filter(
            fecha_solicitud__year=hoy.year,
            fecha_solicitud__month=hoy.month
        )

    elif tiempo == "anio":
        pedidos = pedidos.filter(
            fecha_creacion__year=hoy.year
        )
        devoluciones = devoluciones.filter(
            fecha_solicitud__year=hoy.year
        )

    # FILTROS ESPECÍFICOS
    if mes_especifico:
        pedidos = pedidos.filter(fecha_creacion__month=int(mes_especifico))
        devoluciones = devoluciones.filter(fecha_solicitud__month=int(mes_especifico))
    
    if anio_especifico:
        pedidos = pedidos.filter(fecha_creacion__year=int(anio_especifico))
        devoluciones = devoluciones.filter(fecha_solicitud__year=int(anio_especifico))

    # Obtener los datos
    total_ventas = pedidos.aggregate(total=Sum("total"))["total"] or 0
    total_pedidos = pedidos.count()

    prod_info = (
        Producto.objects
        .annotate(
            total_vendidos=Sum(
                "pedidoitem__cantidad",
                filter=Q(pedidoitem__pedido__in=pedidos)
            )
        )
        .filter(total_vendidos__gt=0)
        .order_by("-total_vendidos")
    )

    usuarios_info = (
        Usuario.objects
        .annotate(
            pedidos_pagados=Count(
                "pedido",
                filter=Q(pedido__in=pedidos)
            )
        )
        .filter(pedidos_pagados__gt=0)
        .order_by("-pedidos_pagados")
    )

    ventas_por_mes = (
        pedidos.annotate(mes=TruncMonth("fecha_creacion"))
        .values("mes")
        .annotate(total=Sum("total"))
        .order_by("mes")
    )

    estado_pedidos = (
        pedidos.values("estado")
        .annotate(total=Count("id"))
        .order_by("estado")
    )

    ultimas_devoluciones = devoluciones.order_by('-fecha_solicitud')[:5]

    # Crear el libro de Excel
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar la hoja por defecto

    # Estilos
    header_fill = PatternFill(start_color="F5A623", end_color="F5A623", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # HOJA 1: Resumen General
    ws1 = wb.create_sheet("Resumen General")
    ws1.append(["DASHBOARD - RESUMEN GENERAL"])
    ws1.merge_cells('A1:B1')
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1.append([])
    ws1.append(["Métrica", "Valor"])
    for cell in ws1[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    ws1.append(["Total Ventas", f"${total_ventas:,.2f}"])
    ws1.append(["Total Pedidos", total_pedidos])
    ws1.append(["Total Productos Vendidos", prod_info.count()])
    ws1.append(["Total Usuarios con Compras", usuarios_info.count()])

    # Ajustar anchos
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20

    # HOJA 2: Productos Más Vendidos
    ws2 = wb.create_sheet("Productos Más Vendidos")
    ws2.append(["Producto", "Cantidad Vendida"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for prod in prod_info:
        ws2.append([prod.nombProduc, prod.total_vendidos])

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 20

    # HOJA 3: Usuarios con Más Compras
    ws3 = wb.create_sheet("Usuarios Top")
    ws3.append(["Usuario", "Cantidad de Pedidos"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for user in usuarios_info:
        ws3.append([user.nombre, user.pedidos_pagados])

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20

    # HOJA 4: Ventas por Mes
    ws4 = wb.create_sheet("Ventas por Mes")
    ws4.append(["Mes", "Total Ventas"])
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for venta in ventas_por_mes:
        ws4.append([venta["mes"].strftime("%B %Y"), f"${venta['total']:,.2f}"])

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 20

    # HOJA 5: Estado de Pedidos
    ws5 = wb.create_sheet("Estado de Pedidos")
    ws5.append(["Estado", "Cantidad"])
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for estado in estado_pedidos:
        ws5.append([estado["estado"], estado["total"]])

    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 20

    # HOJA 6: Últimas Devoluciones
    ws6 = wb.create_sheet("Últimas Devoluciones")
    ws6.append(["Producto", "Unidades", "Motivo", "Estado", "Fecha"])
    for cell in ws6[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for dev in ultimas_devoluciones:
        ws6.append([
            dev.producto.nombProduc,
            dev.unidad,
            dev.motivo,
            dev.estado,
            dev.fecha_solicitud.strftime("%d/%m/%Y")
        ])

    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 50
    ws6.column_dimensions['D'].width = 12
    ws6.column_dimensions['E'].width = 15

    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha
    nombre_archivo = f"Dashboard_{hoy.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    return response