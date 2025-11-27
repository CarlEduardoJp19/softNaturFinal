# Imports de biblioteca estándar de Python
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

# Imports de la aplicación actual (usuarios)
from .models import Devolucion, HistorialDevolucion

# ====================== VISTAS DE DEVOLUCIONES (ADMIN) ======================
@login_required
def gst_devoluciones(request):
    """Vista para que el admin gestione todas las devoluciones"""
    
    # ✅ SOLO VERIFICAR is_staff o is_superuser
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder aquí')
        return redirect('usuarios:dashboard')
    
    estado = request.GET.get('estado')
    
    if estado:
        # Si hay filtro, traer las devoluciones de ese estado
        devoluciones = Devolucion.objects.select_related(
            'usuario', 'producto', 'pedido'
        ).filter(estado=estado).order_by('-fecha_solicitud')
    else:
        # Por defecto, solo las PENDIENTES
        devoluciones = Devolucion.objects.select_related(
            'usuario', 'producto', 'pedido'
        ).filter(estado='Pendiente').order_by('-fecha_solicitud')
    
    # Preparar datos en JSON para el modal
    devoluciones_data = []
    for dev in devoluciones:
        devoluciones_data.append({
            'id': dev.id,
            'usuario_nombre': dev.usuario.nombre,
            'usuario_email': dev.usuario.email,
            'producto_nombre': dev.producto.nombProduc,
            'pedido_id': dev.pedido.id,
            'item_id': dev.item.id if dev.item else None,
            'unidad': dev.unidad,
            'lote': dev.lote.codigo_lote if dev.lote else None, 
            'fecha': dev.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            'motivo': dev.motivo,
            'estado': dev.estado,
            'fotos': dev.get_fotos()
        })
    
    import json
    devoluciones_json = json.dumps(devoluciones_data)
    
    context = {
        'devoluciones': devoluciones,
        'devoluciones_json': devoluciones_json
    }
    
    return render(request, 'usuarios/gst_devoluciones.html', context)

def aprobar_devolucion(request, devolucion_id):
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': "No tienes permisos."}, status=403)
    
    try:
        devolucion = Devolucion.objects.select_related('pedido', 'producto', 'item').get(id=devolucion_id)
        producto = devolucion.producto
        item = devolucion.item
        
        if not item:
            return JsonResponse({'success': False, 'error': "No se encontró el item asociado a la devolución"}, status=400)
        
        # Marcar la devolución como aprobada
        devolucion.estado = "Aprobada"
        devolucion.fecha_respuesta = timezone.now()
        devolucion.save()
        
        # Registrar historial
        comentario_inicial = f"Aprobada. Motivo: {devolucion.motivo}"
        HistorialDevolucion.objects.create(
            devolucion=devolucion,
            estado='Aprobada',
            usuario_admin=request.user,
            comentario=comentario_inicial
        )
        
        # === GESTIÓN DE STOCK DEL PRODUCTO DEVUELTO ===
        lote_original = item.lote
        
        if devolucion.motivo == "Producto equivocado":
            # El producto EQUIVOCADO nunca debió salir del stock (fue error de empaque)
            # El producto CORRECTO ya fue descontado en la compra original
            # Por tanto: NO hay cambios en el stock
            print("\n→ PRODUCTO EQUIVOCADO: Sin cambios en stock")
            print("   - Producto correcto (que pidió) ya fue descontado en compra original")
            print("   - Producto equivocado (que recibió) nunca salió del inventario")
            comentario_stock = "Producto equivocado devuelto. Sin cambios en stock (el correcto ya fue descontado originalmente)"
        
        elif devolucion.motivo in ["Fecha de vencimiento expirado", "Producto dañado"]:
            # El producto vencido/dañado YA FUE DESCONTADO en la compra original
            # El cliente lo tiene en su poder, por tanto NO se descuenta nuevamente
            # Solo se descartará cuando se recoja (pero ya está fuera del inventario)
            print("\n→ PRODUCTO DAÑADO/VENCIDO: Ya fue descontado en compra original")
            print("   - El producto defectuoso está en poder del cliente")
            print("   - NO se descuenta nuevamente (ya salió del inventario)")
            
            if lote_original:
                comentario_stock = f"Producto defectuoso del lote {lote_original.codigo_lote or lote_original.id} (ya descontado en compra)"
            else:
                comentario_stock = "Producto defectuoso (ya descontado en compra)"
        
        else:
            print("\n→ OTRO MOTIVO: Sin cambios en stock")
            comentario_stock = "Sin cambios en el stock"
        
        # Actualizar comentario del historial
        historial = HistorialDevolucion.objects.filter(devolucion=devolucion).last()
        if historial:
            historial.comentario += f" | {comentario_stock}"
            historial.save()
        
        # === GESTIÓN DE REEMPLAZO ===
        mensaje_reemplazo = ""
        lote_reemplazo = None
        
        if devolucion.motivo in ["Fecha de vencimiento expirado", "Producto dañado"]:
            # Buscar un lote disponible para el reemplazo
            lote_reemplazo = producto.lotes.filter(cantidad__gt=0).order_by('fecha_caducidad').first()
            
            if lote_reemplazo:
                # Descontar 1 unidad del lote de reemplazo
                lote_reemplazo.cantidad -= 1
                lote_reemplazo.save()
                print(f"✅ Descontado 1 unidad del lote {lote_reemplazo.codigo_lote or lote_reemplazo.id} (reemplazo)")
                print(f"   Stock lote después: {lote_reemplazo.cantidad}")
                mensaje_reemplazo = f"Se enviará un producto de reemplazo del lote {lote_reemplazo.codigo_lote or lote_reemplazo.id}"
            else:
                print(f"⚠️ ADVERTENCIA: No hay stock disponible para enviar reemplazo")
                mensaje_reemplazo = "⚠️ Sin stock disponible para reemplazo"
                
        elif devolucion.motivo == "Producto equivocado":
            # Buscar un lote disponible del producto CORRECTO
            lote_reemplazo = producto.lotes.filter(cantidad__gt=0).order_by('fecha_caducidad').first()
            
            if lote_reemplazo:
                print(f"✅ Se enviará el producto correcto del lote {lote_reemplazo.codigo_lote or lote_reemplazo.id} (sin descuento adicional)")
                mensaje_reemplazo = f"Se enviará el producto correcto del lote {lote_reemplazo.codigo_lote or lote_reemplazo.id}"
            else:
                print(f"⚠️ ADVERTENCIA: No hay lotes disponibles del producto correcto")
                mensaje_reemplazo = "⚠️ Sin stock disponible del producto correcto"
        else:
            mensaje_reemplazo = "Se gestionará el envío del reemplazo"
        
        # === ENVIAR EMAIL AL CLIENTE ===
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            # Obtener nombre del usuario
            nombre_usuario = devolucion.usuario.nombre if devolucion.usuario.nombre else devolucion.usuario.email.split('@')[0]
            
            asunto = f'Devolución #{devolucion.id} aprobada - {producto.nombProduc if hasattr(producto, "nombProduc") else str(producto)}'
            
            mensaje = f"""
Hola {nombre_usuario},

Tu devolución del producto "{producto.nombProduc if hasattr(producto, "nombProduc") else str(producto)}" ha sido APROBADA.

DETALLES DE LA DEVOLUCIÓN:
- Devolución #: {devolucion.id}
- Producto: {producto.nombProduc if hasattr(producto, "nombProduc") else str(producto)}
- Unidad: {devolucion.unidad}
- Motivo: {devolucion.motivo}
- Fecha de aprobación: {devolucion.fecha_respuesta.strftime('%d/%m/%Y %H:%M')}

REEMPLAZO:
{mensaje_reemplazo}

Pronto recibirás tu producto de reemplazo{f' (Lote: {lote_reemplazo.codigo_lote or lote_reemplazo.id}, Vencimiento: {lote_reemplazo.fecha_caducidad.strftime("%d/%m/%Y")})' if lote_reemplazo else ''}.

Gracias por tu confianza.

---
Este es un correo automático.
            """.strip()
            
            # Verificar que el usuario tenga email
            if devolucion.usuario.email:
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[devolucion.usuario.email],
                    fail_silently=True
                )
            else:
                print(f"Usuario {devolucion.usuario.nombre} no tiene email configurado")
            
        except Exception as e:
            print(f"Error al enviar email: {str(e)}")
        
        # Respuesta AJAX final
        return JsonResponse({
            'success': True,
            'id': devolucion_id,
            'mensaje': f"Devolución #{devolucion_id} aprobada. {mensaje_reemplazo}"
        })
        
    except Devolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Devolución no encontrada'}, status=404)
    except Exception as e:
        print(f"Error en aprobar_devolucion: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@login_required
def rechazar_devolucion(request, devolucion_id):
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': "No tienes permisos."}, status=403)

    try:
        devolucion = Devolucion.objects.get(id=devolucion_id)
        devolucion.estado = 'Rechazada'
        devolucion.fecha_respuesta = timezone.now()
        devolucion.save()

        HistorialDevolucion.objects.create(
            devolucion=devolucion,
            estado='Rechazada',
            usuario_admin=request.user,
            comentario='Rechazada por admin'
        )

        return JsonResponse({
            'success': True,
            'id': devolucion_id,
            'mensaje': f"Devolución #{devolucion_id} rechazada"
        })

    except Devolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Devolución no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
def historial_devoluciones(request):
    """Vista para mostrar el historial de devoluciones (solo Aprobadas o Rechazadas)."""
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, "No tienes permisos para acceder aquí")
        return redirect("usuarios:dashboard")
    
    filtro_tiempo = request.GET.get('tiempo')
    estado = request.GET.get('estado')
    busqueda = request.GET.get('buscar')
    
    ahora = timezone.now()
    
    # Base query - solo Aprobadas o Rechazadas
    devoluciones = Devolucion.objects.filter(
        estado__in=['Aprobada', 'Rechazada']
    ).select_related('usuario', 'producto', 'pedido', 'lote').prefetch_related('historial')
    
    # Filtro por tiempo
    if filtro_tiempo == "semana":
        inicio_semana = ahora - timezone.timedelta(days=ahora.weekday())
        devoluciones = devoluciones.filter(fecha_solicitud__gte=inicio_semana)
    elif filtro_tiempo == "mes":
        devoluciones = devoluciones.filter(
            fecha_solicitud__year=ahora.year,
            fecha_solicitud__month=ahora.month
        )
    elif filtro_tiempo == "año":
        devoluciones = devoluciones.filter(
            fecha_solicitud__year=ahora.year
        )
    
    # Filtro por estado (Aprobada o Rechazada)
    if estado:
        devoluciones = devoluciones.filter(estado=estado)
    
    # Filtro por búsqueda (usuario, producto o lote)
    if busqueda:
        devoluciones = devoluciones.filter(
            Q(usuario__nombre__icontains=busqueda) |
            Q(usuario__email__icontains=busqueda) |
            Q(producto__nombProduc__icontains=busqueda) |
            Q(lote__codigo_lote__icontains=busqueda)
        )
    
    # Orden
    devoluciones = devoluciones.order_by('-fecha_solicitud')
    
    # Conteos para estadísticas
    total_aprobadas = devoluciones.filter(estado='Aprobada').count()
    total_rechazadas = devoluciones.filter(estado='Rechazada').count()
    
    context = {
        'devoluciones': devoluciones,
        'total_aprobadas': total_aprobadas,
        'total_rechazadas': total_rechazadas,
        'filtro_tiempo': filtro_tiempo,
        'estado_filtro': estado,
        'busqueda': busqueda,
    }
    return render(request, 'usuarios/historial_devoluciones.html', context)

def exportar_devoluciones_excel(request):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Devoluciones"

    # Encabezado principal (NARANJA)
    ws.merge_cells('A1:H1')
    ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    ws.row_dimensions[1].height = 30

    # Subtítulo (NARANJA OSCURO)
    ws.merge_cells('A2:H2')
    ws['A2'] = 'HISTORIAL DE DEVOLUCIONES'
    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    ws.row_dimensions[2].height = 25

    # Fecha de generación
    ws.merge_cells('A3:H3')
    ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Fila vacía
    ws.append([])

    # Encabezados de columnas (AMARILLO)
    encabezados = ['ID', 'Cliente', 'Email', 'Producto', 'Lote', 'Motivo', 'Estado', 'Fecha']
    ws.append(encabezados)

    header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, 9):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Filtro desde GET
    estado = request.GET.get('estado')

    # Consulta base
    devoluciones = Devolucion.objects.filter(
        estado__in=['Aprobada', 'Rechazada']
    ).order_by('-fecha_solicitud')
    
    if estado:
        devoluciones = devoluciones.filter(estado=estado)

    # Colores por estado
    estado_colors = {
        'aprobada': 'D5F4E6',   # Verde claro
        'rechazada': 'FADBD8',  # Rojo claro
    }

    # Datos
    row_num = 6
    for dev in devoluciones:
        ws.append([
            dev.id,
            dev.usuario.nombre if dev.usuario else 'Sin usuario',
            dev.usuario.email if dev.usuario else 'Sin email',
            dev.producto.nombProduc if dev.producto else 'Sin producto',
            dev.lote.codigo_lote if dev.lote else 'Sin lote',
            dev.motivo,
            dev.estado,
            dev.fecha_solicitud.strftime('%d/%m/%Y %H:%M')
        ])

        # Color según estado
        estado_lower = dev.estado.lower()
        row_fill = PatternFill(
            start_color=estado_colors.get(estado_lower, 'FFFFFF'),
            end_color=estado_colors.get(estado_lower, 'FFFFFF'),
            fill_type='solid'
        )

        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.fill = row_fill
        row_num += 1

    # Resumen al final
    total_aprobadas = devoluciones.filter(estado='Aprobada').count()
    total_rechazadas = devoluciones.filter(estado='Rechazada').count()
    
    ws.append([])
    row_num += 1
    
    ws.append(['', '', '', '', '', 'Total Aprobadas:', total_aprobadas, ''])
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)
        if col == 7:
            cell.fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
    row_num += 1

    ws.append(['', '', '', '', '', 'Total Rechazadas:', total_rechazadas, ''])
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)
        if col == 7:
            cell.fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')

    # Ajustar ancho de columnas
    column_widths = [10, 25, 30, 25, 15, 35, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_devoluciones.xlsx"'
    wb.save(response)
    return response

