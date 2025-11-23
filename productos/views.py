from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Q
from django.contrib import messages
import openpyxl
from django.utils import timezone
from datetime import timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import os
import google.generativeai as genai
import json
from .forms import ProductoForm, CategoriaForm
from .models import Producto, Category, Servicio, Calificacion, CarritoItem, Lote
from usuarios.models import Usuario, Pedido, PedidoItem
from usuarios.decorators import admin_required
from usuarios.models import Devolucion, Pedido  # Importar de usuarios
from usuarios.decorators import login_required
from datetime import date



# ---------------------- VISTAS DE PRODUCTOS ----------------------

def productos(request):
    products = Producto.objects.all()
    return render(request, 'productos/producto.html', {'products': products})

@admin_required
def list_product(request):
    # Obtener productos con sus lotes (optimizado)
    productos = Producto.objects.prefetch_related('lotes').all()
    
    # Obtener la fecha actual
    today = timezone.now().date()
    
    # Procesar cada producto para agregar datos adicionales
    for producto in productos:
        # Lote más próximo a vencer (para mostrar en la tarjeta)
        lote_proximo = producto.lotes.order_by('fecha_caducidad').first()
        producto.lote_mas_proximo = lote_proximo
        
        # ✅ AÑADIR ESTO: Lote activo (el que se vende primero - con stock y más próximo a vencer)
        producto.lote_activo = producto.lotes.filter(
            cantidad__gt=0
        ).order_by('fecha_caducidad').first()
        
        # Total de lotes (para mostrar en template)
        producto.total_lotes = producto.stock_total
    
    # Obtener categorías
    categorias = Category.objects.all()
    
    from .forms import ProductoForm
    form = ProductoForm()
    
    return render(request, 'productos/list_produc.html', {
        'productos': productos,
        'categorias': categorias,
        'today': today,
        'form': form
    })


def productos_view(request, categoria_id=None):
    buscar = request.GET.get('buscar', '')

    if categoria_id:
        try:
            categoria = Category.objects.get(id=categoria_id)
            productos = Producto.objects.filter(categoria=categoria, estado=True)
        except Category.DoesNotExist:
            productos = Producto.objects.filter(estado=True)
            categoria = None
    else:
        productos = Producto.objects.filter(estado=True)
        categoria = None

    if buscar:
        productos = productos.filter(
            Q(nombProduc__icontains=buscar) |
            Q(descripcion__icontains=buscar)
        )

    # ------------------ Carrito en sesión ------------------
    carrito = request.session.get('carrito', {})
    items = []
    subtotal = 0
    total_cantidad = 0

    for key, value in carrito.items():
        if not isinstance(value, dict):
            continue

        cantidad = int(value.get("cantidad", 0))
        precio = float(value.get("precio", 0))
        nombre = value.get("nombProduc", "Sin nombre")
        precio_total = cantidad * precio

        subtotal += precio_total
        total_cantidad += cantidad

        items.append({
            "producto_id": key,
            "producto": nombre,
            "cantidad": cantidad,
            "precio_unitario": precio,
            "precio_total": precio_total
        })

    iva = subtotal * 0.19
    total = subtotal + iva

    categorias = Category.objects.all()

    return render(request, 'productos/producto.html', {
        'products': productos,
        'carrito': items,
        'subtotal': round(subtotal, 2),
        'iva': round(iva, 2),
        'total': round(total, 2),
        'total_cantidad': total_cantidad,
        'categorias': categorias,
        'categoria_actual': categoria,
        'buscar': buscar,
    })

# ---------------------- CARRITO ----------------------
def agregarAlCarrito(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    carrito = request.session.get("carrito", {})
    producto_id_str = str(producto.id)

    # Buscar el lote más cercano a vencer
    lote = Lote.objects.filter(
        producto=producto,
        cantidad__gt=0
    ).order_by('fecha_caducidad').first()

    # 🔍 DEBUG DETALLADO
    print("=" * 50)
    print(f"PRODUCTO: {producto.id} - {producto.nombProduc}")
    print(f"LOTE ENCONTRADO: {lote}")
    if lote:
        print(f"  → codigo_lote: {lote.codigo_lote}")
        print(f"  → cantidad: {lote.cantidad}")
        print(f"  → producto_id del lote: {lote.producto_id}")
    else:
        print("  → ❌ NO SE ENCONTRÓ NINGÚN LOTE")
        
        # Ver qué lotes existen para este producto
        todos_lotes = Lote.objects.filter(producto=producto)
        print(f"  → Lotes totales para este producto: {todos_lotes.count()}")
        for l in todos_lotes:
            print(f"     - {l.codigo_lote}, cantidad: {l.cantidad}")
    print("=" * 50)

    lote_codigo = lote.codigo_lote if lote else None

    if producto_id_str in carrito:
        carrito[producto_id_str]["cantidad"] += 1
        if not carrito[producto_id_str].get("lote"):
            carrito[producto_id_str]["lote"] = lote_codigo
    else:
        carrito[producto_id_str] = {
            "cantidad": 1,
            "precio": float(producto.precio),
            "nombProduc": producto.nombProduc,
            "imgProduc": producto.imgProduc.url,
            "lote": lote_codigo
        }

    request.session["carrito"] = carrito
    request.session.modified = True

    return redirect(f"{reverse('productos:producto')}?carrito=1")



def eliminar(request, producto_id):
    carrito = request.session.get('carrito', {})
    producto_id_str = str(producto_id)

    # --- SESIÓN ---
    if producto_id_str in carrito:
        del carrito[producto_id_str]

    request.session['carrito'] = carrito
    request.session.modified = True

    # --- BASE DE DATOS ---
    if request.user.is_authenticated:
        CarritoItem.objects.filter(usuario=request.user, producto_id=producto_id).delete()

    return redirect("productos:producto")

def restar_producto(request, producto_id):
    carrito = request.session.get('carrito', {})
    producto_id_str = str(producto_id)

    # --- SESIÓN ---
    if producto_id_str in carrito:
        if carrito[producto_id_str]["cantidad"] > 1:
            carrito[producto_id_str]["cantidad"] -= 1
        else:
            del carrito[producto_id_str]

    request.session['carrito'] = carrito
    request.session.modified = True

    # --- BASE DE DATOS ---
    if request.user.is_authenticated:
        try:
            item = CarritoItem.objects.get(usuario=request.user, producto_id=producto_id)
            if item.cantidad > 1:
                item.cantidad -= 1
                item.save()
            else:
                item.delete()
        except CarritoItem.DoesNotExist:
            pass

    return redirect(f"{reverse('productos:producto')}?carrito=1")

def limpiar(request):
    # --- SESIÓN ---
    request.session['carrito'] = {}
    request.session.modified = True

    # --- BASE DE DATOS ---
    if request.user.is_authenticated:
        CarritoItem.objects.filter(usuario=request.user).delete()

    return redirect("productos:producto")

# ---------------------- CRUD PRODUCTOS ----------------------
@admin_required
def agregar_producto(request):
    form = ProductoForm(request.POST or None, request.FILES or None)
    productos = Producto.objects.all()

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Producto creado con éxito.")
            return redirect('productos:agregar_producto')
        else:
            messages.error(request, "Errores en el formulario. Revisa los campos.")

    return render(request, 'productos/list_produc.html', {
        'form': form,
        'productos': productos
    })

@admin_required
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('productos:list_product')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'productos/editar_producto.html', {'form': form, 'producto': producto})

def productos_por_categoria(request, categoria_id):
    try:
        categoria = Category.objects.get(id=categoria_id)
        productos = Producto.objects.filter(Categoria=categoria, estado=True)
    except Category.DoesNotExist:
        productos = Producto.objects.filter(estado=True)
        categoria = None

    usuario_id = request.session.get('usuario_id')
    carrito = request.session.get('carrito', {})
    items = []
    total = 0

    for key, value in carrito.items():
        if not isinstance(value, dict):
            continue

        cantidad = int(value.get("cantidad", 0))
        precio = float(value.get("precio", 0))
        nombre = value.get("nombProduc", "Sin nombre")

        subtotal = precio * cantidad
        items.append({
            "producto": nombre,
            "cantidad": cantidad,
            "precio": precio,
            "subtotal": subtotal
        })
        total += subtotal

    categorias = Category.objects.all()

    return render(request, 'productos/producto.html', {
        'products': productos,
        'usuario_id': usuario_id,
        'carrito': items,
        'total': total,
        'categorias': categorias,
        'categoria_actual': categoria,
    })

# ---------------------- INVENTARIO / EXPORTAR ----------------------

@admin_required
def exportar_inventario_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    # Encabezado principal de la tienda (NARANJA)
    ws.merge_cells('A1:I1')
    ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    ws.row_dimensions[1].height = 30

    # Subtítulo (NARANJA OSCURO)
    ws.merge_cells('A2:I2')
    ws['A2'] = 'INVENTARIO DE PRODUCTOS'
    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    ws.row_dimensions[2].height = 25

    # Fecha de generación
    ws.merge_cells('A3:I3')
    ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Fila vacía
    ws.append([])

    # Encabezados de columnas (fila 5)
    encabezados = [
        'Nombre', 
        'Descripción', 
        'Precio', 
        'Stock Total', 
        'Categoría',
        'Código Lote',
        'Cantidad Lote',
        'Fecha Caducidad',
        'Lote Activo'
    ]
    ws.append(encabezados)

    # Estilo para encabezados de columnas (AMARILLO)
    header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, 10):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Datos de productos
    productos = Producto.objects.prefetch_related('lotes').all()
    row_num = 6

    # Color para filas de lote activo (AMARILLO CLARO)
    lote_activo_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')

    for producto in productos:
        lote_activo = producto.lotes.filter(
            cantidad__gt=0
        ).order_by('fecha_caducidad').first()
        
        lotes = producto.lotes.all()
        
        if lotes.exists():
            for lote in lotes:
                es_activo = "Sí" if lote_activo and lote.id == lote_activo.id else "No"
                ws.append([
                    producto.nombProduc,
                    producto.descripcion,
                    float(producto.precio),
                    producto.stock_total,
                    str(producto.Categoria),
                    lote.codigo_lote,
                    lote.cantidad,
                    lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'Sin fecha',
                    es_activo
                ])
                
                for col in range(1, 10):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = thin_border
                    if es_activo == "Sí":
                        cell.fill = lote_activo_fill
                
                row_num += 1
        else:
            ws.append([
                producto.nombProduc,
                producto.descripcion,
                float(producto.precio),
                producto.stock_total,
                str(producto.Categoria),
                'Sin lote',
                0,
                'N/A',
                'N/A'
            ])
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

    # Ajustar ancho de columnas
    column_widths = [25, 40, 12, 12, 20, 15, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Inventario_de_productos.xlsx'
    wb.save(response)
    return response

@admin_required
def activar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.estado = True
    producto.save()
    return redirect('productos:listProduc')

@admin_required
def inactivar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.estado = False
    producto.save()
    return redirect('productos:listProduc')

# ---------------------- CALIFICACIONES ----------------------
def guardar_calificacion(request):
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')   # ✅ viene del formulario
        puntuacion_servicio = request.POST.get('puntuacion_servicio')
        puntuacion_productos = request.POST.get('puntuacion_productos')
        comentario = request.POST.get('comentario')

        print("Datos recibidos:")
        print("producto_id:", producto_id)
        print("puntuacion_servicio:", puntuacion_servicio)
        print("puntuacion_productos:", puntuacion_productos)
        print("comentario:", comentario)

        try:
            producto = Producto.objects.get(id=producto_id)  # ✅ ya no busca Servicio
        except Producto.DoesNotExist:
            return HttpResponse("❌ Producto no encontrado", status=404)

        usuario_obj = request.user

        Calificacion.objects.create(
            producto=producto,                 # ✅ guarda asociado al producto
            usuario=usuario_obj,
            puntuacion_servicio=int(puntuacion_servicio),
            puntuacion_productos=int(puntuacion_productos),
            comentario=comentario
        )

        print("✅ Calificación guardada correctamente")

        return render(request, 'productos/homeSoft.html')

    return HttpResponse("❌ Método no permitido", status=405)

# ---------------------- CARGAR CARRITO USUARIO ----------------------

def cargar_carrito_usuario(request, usuario):
    carrito_items = CarritoItem.objects.filter(usuario=usuario)
    carrito = {}

    for item in carrito_items:
        carrito[str(item.producto.id)] = {
            "cantidad": item.cantidad,
            "precio": float(item.producto.precio),
            "nombProduc": item.producto.nombProduc,
        }

    request.session["carrito"] = carrito
    request.session.modified = True

# ---------------------- CRUD CATEGORÍAS ----------------------

@admin_required
def agregar_categoria(request):
    if request.method == "POST":
        nombre = request.POST.get("nombCategory")
        if nombre:
            Category.objects.create(nombCategory=nombre)
    return redirect('productos:listar_categorias')


@admin_required
def editar_categoria(request, id):
    categoria = get_object_or_404(Category, id=id)
    if request.method == "POST":
        nuevo_nombre = request.POST.get("nombre")
        if nuevo_nombre:
            categoria.nombCategory = nuevo_nombre
            categoria.save()
    return redirect('productos:listar_categorias')


@admin_required
def listar_categorias(request):
    categorias = Category.objects.all()
    productos = Producto.objects.all()
    form = CategoriaForm()

    if request.method == "POST":
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('productos:listar_categorias')

    return render(request, 'productos/list_produc.html', {
        'categorias': categorias,
        'productos': productos,
        'form': form
    })

@admin_required
def eliminar_categoria(request, id):
    categoria = get_object_or_404(Category, id=id)
    categoria.delete()
    return redirect('productos:listar_categorias')

# ---------------------- HOME ----------------------

def homeSoft(request):
    mas_vendidos = Producto.objects.filter(estado=True).order_by("vendidos")[:3]
    comentarios = Calificacion.objects.filter(aprobado=True).order_by('-fecha_creacion')
    return render(request, "productos/homeSoft.html", {
        "mas_vendidos": mas_vendidos,
        "comentarios": comentarios
    })

# ========== IMPORTS PARA VISTA DE DEVOLUCIONES ==========                                                                                                                                     
@login_required
def devoluciones(request):
    """Vista para que el cliente solicite devoluciones por unidad"""
    # Cambiar de 30 días a 10 días
    hace_10_dias = timezone.now() - timedelta(days=10)
    pedidos = Pedido.objects.filter(
        usuario=request.user,
        estado='entregado',
        fecha_creacion__gte=hace_10_dias
    ).prefetch_related('items__producto').order_by('-fecha_creacion')

    # Devoluciones existentes (cualquier estado, no solo Pendiente)
    # Esto evita que se pueda devolver un producto ya devuelto
    devoluciones_existentes = Devolucion.objects.filter(
        usuario=request.user
    ).values_list('producto_id', 'pedido_id', 'unidad')
    devoluciones_existentes = [(int(p), int(d), int(u)) for p, d, u in devoluciones_existentes]

    productos_devolubles = []
    pedidos_agrupados = {}

    for pedido in pedidos:
        # Verificar que el pedido tenga menos de 10 días
        dias_transcurridos = (timezone.now() - pedido.fecha_creacion).days
        if dias_transcurridos > 10:
            continue  # Saltar pedidos con más de 10 días
        
        items = pedido.items.all()
        unidades_disponibles = 0
        
        for item in items:
            cantidad = getattr(item, 'cantidad', 1) or 1
            for unidad_index in range(cantidad):
                unidad_num = unidad_index + 1
                key = (getattr(item.producto, 'id', None), pedido.id, unidad_num)
                
                # Si ya existe una devolución para esta unidad, no mostrarla
                if key in devoluciones_existentes:
                    continue
                
                unidades_disponibles += 1

                # Normalizar lote a string para JSON
                lote_val = getattr(item, 'lote', None)
                if lote_val is None:
                    lote_str = ''
                elif isinstance(lote_val, str):
                    lote_str = lote_val
                else:
                    lote_str = getattr(lote_val, 'codigo_lote', None) or getattr(lote_val, 'codigo', None) or str(lote_val)

                productos_devolubles.append({
                    'pedido_id': pedido.id,
                    'producto_id': getattr(item.producto, 'id', None),
                    'producto_nombre': getattr(item.producto, 'nombProduc', str(item.producto)),
                    'unidad': unidad_num,
                    'precio': float(getattr(item, 'precio_unitario', 0) or 0),
                    'item_id': item.id,
                    'lote': lote_str,
                    'codigo_lote': lote_str,
                    'fecha_pedido': pedido.fecha_creacion.strftime('%d/%m/%Y') if getattr(pedido, 'fecha_creacion', None) else ''
                })

        if unidades_disponibles > 0:
            pedidos_agrupados[pedido.id] = {
                'pedido_id': pedido.id,
                'fecha_pedido': pedido.fecha_creacion.strftime('%d/%m/%Y'),
                'cantidad_productos': unidades_disponibles,
                'dias_restantes': 10 - dias_transcurridos  # Días que le quedan para devolver
            }

    # POST: crear devolución
    if request.method == 'POST':
        try:
            pedido_id = int(request.POST.get('pedido_id'))
            motivo = request.POST.get('motivo')
            foto1 = request.FILES.get('foto1')
            foto2 = request.FILES.get('foto2')
            foto3 = request.FILES.get('foto3')
            
            # Procesar producto_data
            producto_data = request.POST.get('producto_id')
            
            if '|' in producto_data:
                # Formato: producto_id|item_id|lote_codigo
                datos = producto_data.split('|')
                producto_id = int(datos[0])
                item_id = int(datos[1]) if len(datos) > 1 and datos[1] and datos[1] != 'None' else None
                lote_codigo = datos[2] if len(datos) > 2 else ''
            elif '-' in producto_data:
                # Formato antiguo: producto_id-unidad
                producto_id, _ = map(int, producto_data.split('-'))
                item_id = None
                lote_codigo = ''
            else:
                producto_id = int(producto_data)
                item_id = None
                lote_codigo = ''
            
            # Obtener unidad del POST
            unidad = int(request.POST.get('unidad', 1))

            # Validar campos obligatorios
            if not producto_id or not pedido_id or not motivo or not unidad:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({"success": False, "mensaje": "Completa todos los campos"})
                messages.error(request, "Por favor completa todos los campos obligatorios")
                return redirect('productos:devoluciones')

            # Obtener producto y pedido
            producto = Producto.objects.get(id=producto_id)
            pedido = Pedido.objects.get(id=pedido_id, usuario=request.user, estado='entregado')
            
            # Buscar el pedido_item con mejor precisión
            pedido_item = None
            
            # Opción 1: Si tenemos item_id, usarlo directamente
            if item_id:
                pedido_item = pedido.items.filter(id=item_id, producto=producto).first()
            
            # Opción 2: Si no hay item_id, buscar por producto
            if not pedido_item:
                items_producto = pedido.items.filter(producto=producto)
                
                # Si solo hay un item con ese producto, usarlo
                if items_producto.count() == 1:
                    pedido_item = items_producto.first()
                # Si hay múltiples, intentar filtrar por lote si está disponible
                elif items_producto.count() > 1 and lote_codigo:
                    # Intentar filtrar por lote
                    for item in items_producto:
                        item_lote = getattr(item, 'lote', None)
                        if item_lote:
                            if isinstance(item_lote, str):
                                if item_lote == lote_codigo:
                                    pedido_item = item
                                    break
                            else:
                                lote_code = getattr(item_lote, 'codigo_lote', None) or getattr(item_lote, 'codigo', None)
                                if lote_code == lote_codigo:
                                    pedido_item = item
                                    break
                    
                    # Si no se encontró por lote, tomar el primero
                    if not pedido_item:
                        pedido_item = items_producto.first()
                else:
                    # Tomar el primero disponible
                    pedido_item = items_producto.first()

            if not pedido_item:
                error_msg = f"No se encontró el producto en el pedido. Producto ID: {producto_id}, Pedido ID: {pedido_id}, Item ID: {item_id}"
                print(error_msg)  # Para debug en logs
                
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({"success": False, "mensaje": "No se encontró el producto en el pedido"})
                messages.error(request, "No se encontró el producto en el pedido")
                return redirect('productos:devoluciones')

            # Verificar cantidad disponible en el item
            cantidad_item = getattr(pedido_item, 'cantidad', 1) or 1
            if unidad > cantidad_item:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({"success": False, "mensaje": f"La unidad {unidad} no existe en este producto (máx: {cantidad_item})"})
                messages.error(request, f"La unidad {unidad} no existe en este producto")
                return redirect('productos:devoluciones')

            # Verificar devolución existente (cualquier estado)
            devolucion_existente = Devolucion.objects.filter(
                usuario=request.user,
                producto=producto,
                pedido=pedido,
                unidad=unidad
            ).exists()

            if devolucion_existente:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({"success": False, "mensaje": "Ya existe una devolución para esta unidad"})
                messages.warning(request, "Ya existe una solicitud de devolución para esta unidad.")
                return redirect('productos:devoluciones')
            
            # Verificar que el pedido tenga menos de 10 días
            dias_transcurridos = (timezone.now() - pedido.fecha_creacion).days
            if dias_transcurridos > 10:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({"success": False, "mensaje": "Este pedido tiene más de 10 días. Ya no puedes solicitar devoluciones."})
                messages.error(request, "Este pedido tiene más de 10 días. Ya no puedes solicitar devoluciones.")
                return redirect('productos:devoluciones')

            # Obtener el lote del pedido_item
            lote_obj = getattr(pedido_item, 'lote', None)

            # Crear devolución
            devolucion = Devolucion(
                usuario=request.user,
                producto=producto,
                pedido=pedido,
                item=pedido_item,
                lote=lote_obj,
                motivo=motivo,
                estado='Pendiente',
                unidad=unidad,
                seleccionada=True
            )

            if foto1:
                devolucion.foto1 = foto1
            if foto2:
                devolucion.foto2 = foto2
            if foto3:
                devolucion.foto3 = foto3

            devolucion.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    "success": True,
                    "producto_id": producto.id,
                    "pedido_id": pedido.id,
                    "unidad": unidad,
                    "mensaje": f"Devolución #{devolucion.id} enviada exitosamente"
                })

            messages.success(request, f"✅ Solicitud de devolución #{devolucion.id} enviada exitosamente!")
            return redirect('productos:devoluciones')

        except Producto.DoesNotExist:
            error_msg = "Producto no encontrado"
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "mensaje": error_msg})
            messages.error(request, error_msg)
            return redirect('productos:devoluciones')
            
        except Pedido.DoesNotExist:
            error_msg = "Pedido no encontrado o no autorizado"
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "mensaje": error_msg})
            messages.error(request, error_msg)
            return redirect('productos:devoluciones')
            
        except ValueError as e:
            error_msg = f"Error en los datos enviados: {str(e)}"
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "mensaje": error_msg})
            messages.error(request, error_msg)
            return redirect('productos:devoluciones')
            
        except Exception as e:
            error_msg = f"Error inesperado: {str(e)}"
            print(f"Error en devoluciones: {error_msg}")  # Para logs
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "mensaje": "Ocurrió un error al procesar la devolución"})
            messages.error(request, error_msg)
            return redirect('productos:devoluciones')

    # Serializar JSON
    productos_json = json.dumps(productos_devolubles, default=str, ensure_ascii=False)
    mis_devoluciones = Devolucion.objects.filter(
        usuario=request.user
    ).select_related('producto', 'pedido').order_by('-fecha_solicitud')
    
    # Timestamp para cache busting
    import time
    timestamp = str(int(time.time()))

    context = {
        'productos_devolubles': productos_json,
        'pedidos_agrupados': pedidos_agrupados.values(),
        'devoluciones': mis_devoluciones,
        'timestamp': timestamp  # Para evitar caché
    }

    return render(request, 'productos/devoluciones.html', context)# Configurar la API Key de Gemini
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

@csrf_exempt
def chatbot_ajax(request):
    if request.method == "POST":
        pregunta = request.POST.get("pregunta", "").strip()
        if not pregunta:
            return JsonResponse({"respuesta": "No se recibió ninguna pregunta."})

        # Prompt restrictivo para Gemini
        prompt = (
            "Eres un chatbot experto únicamente en productos naturistas, suplementos y hierbas. "
            "Si la pregunta del usuario NO está relacionada con productos naturistas, responde: "
            "'Lo siento, solo puedo responder preguntas sobre productos naturistas.' "
            f"Usuario pregunta: {pregunta}. Responde de forma amigable y clara."
        )

        try:
            # Llamada a Gemini
            model = genai.GenerativeModel("models/gemini-pro-latest")
            respuesta = model.generate_content(prompt).text.strip()

            # Fallback si Gemini no devuelve nada
            if not respuesta:
                respuesta = (
                    "Lo siento, solo puedo responder preguntas sobre productos naturistas. "
                    "Por ejemplo, puedes preguntarme sobre hierbas, vitaminas o suplementos naturales."
                )

        except Exception:
            respuesta = (
                "Ocurrió un error al procesar tu pregunta. "
                "Recuerda que solo puedo responder sobre productos naturistas, como hierbas, vitaminas o suplementos."
            )

        return JsonResponse({"respuesta": respuesta})

    return JsonResponse({"error": "Método no permitido"}, status=405)

def detalle_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, "productos/detalle.html", {"producto": producto})

def agregar_lote(request):
    if request.method == "POST":
        producto_id = request.POST.get("producto_id")
        codigo_lote = request.POST.get("codigo_lote")
        fecha_caducidad = request.POST.get("fecha_caducidad")
        cantidad = request.POST.get("cantidad")

        producto = get_object_or_404(Producto, pk=producto_id)

        Lote.objects.create(
            producto=producto,
            codigo_lote=codigo_lote,
            fecha_caducidad=fecha_caducidad,
            cantidad=cantidad
        )

        return redirect("productos:list_product")  

    return redirect("productos:list_product")

def lote_activo(request, producto_id):
    product = get_object_or_404(Producto, id=producto_id)

    # Obtener el lote más próximo a vencer (sin vencer)
    lote_act = product.lotes.filter(
        fecha_caducidad__gte=timezone.now().date()
    ).order_by('fecha_caducidad').first()

    return render(request, 'productos/detalle.html', {
        'product': product,
        'lote_activo': lote_act,   # <--- nombre correcto para el template
    })