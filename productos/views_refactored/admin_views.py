from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponse
from django.contrib import messages
import openpyxl
from django.utils import timezone
from datetime import datetime

# Imports de modelos y forms
from ..forms import ProductoForm, CategoriaForm
from ..models import Producto, Category, Lote

# Decoradores
from usuarios.decorators import admin_required

# Para Excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ==================== LISTADO DE PRODUCTOS ====================

@admin_required
def list_product(request):
    """Vista principal del administrador para gestionar productos y lotes"""
    # Obtener productos con sus lotes (optimizado)
    productos = Producto.objects.prefetch_related('lotes').all()
    
    # Obtener la fecha actual
    today = timezone.now().date()
    
    # Procesar cada producto para agregar datos adicionales
    for producto in productos:
        # Lote más próximo a vencer (para mostrar en la tarjeta)
        lote_proximo = producto.lotes.order_by('fecha_caducidad').first()
        producto.lote_mas_proximo = lote_proximo
        
        # Lote activo (el que se vende primero - con stock y más próximo a vencer)
        producto.lote_activo = producto.lotes.filter(
            cantidad__gt=0
        ).order_by('fecha_caducidad').first()
        
        # Total de lotes
        producto.total_lotes = producto.stock_total
    
    # Obtener categorías
    categorias = Category.objects.all()
    
    form = ProductoForm()
    
    return render(request, 'productos/list_produc.html', {
        'productos': productos,
        'categorias': categorias,
        'today': today,
        'form': form
    })


# ==================== CRUD PRODUCTOS ====================

@admin_required
def agregar_producto(request):
    form = ProductoForm(request.POST or None, request.FILES or None)
    productos = Producto.objects.all()

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Producto creado con éxito.")
            return redirect('productos:agregar_producto')
        else:
            messages.error(request, "Errores en el formulario. Revisa los campos.")

    return render(request, 'productos/list_produc.html', {
        'form': form,
        'productos': productos
    })


@admin_required
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('productos:list_product')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'productos/editar_producto.html', {'form': form, 'producto': producto})


@admin_required
def activar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.estado = True
    producto.save()
    return redirect('productos:list_product')  # ✅ Corregido


@admin_required
def inactivar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.estado = False
    producto.save()
    return redirect('productos:list_product')  # ✅ Corregido


# ==================== GESTIÓN DE LOTES ====================

@admin_required
def agregar_lote(request):
    if request.method == "POST":
        producto_id = request.POST.get("producto_id")
        codigo_lote = request.POST.get("codigo_lote")
        fecha_caducidad = request.POST.get("fecha_caducidad")
        cantidad = request.POST.get("cantidad")

        producto = get_object_or_404(Producto, pk=producto_id)

        Lote.objects.create(
            producto=producto,
            codigo_lote=codigo_lote,
            fecha_caducidad=fecha_caducidad,
            cantidad=cantidad
        )

        return redirect("productos:list_product")
    
    return redirect("productos:list_product")


# ==================== EXPORTAR INVENTARIO A EXCEL ====================

@admin_required
def exportar_inventario_excel(request):
    """Exporta el inventario completo de productos y lotes a Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    # Encabezado principal de la tienda (NARANJA)
    ws.merge_cells('A1:I1')
    ws['A1'] = 'TIENDA NATURISTA LOS GIRASOLES'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
    ws.row_dimensions[1].height = 30

    # Subtítulo (NARANJA OSCURO)
    ws.merge_cells('A2:I2')
    ws['A2'] = 'INVENTARIO DE PRODUCTOS'
    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    ws.row_dimensions[2].height = 25

    # Fecha de generación
    ws.merge_cells('A3:I3')
    ws['A3'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Fila vacía
    ws.append([])

    # Encabezados de columnas (fila 5)
    encabezados = [
        'Nombre', 
        'Descripción', 
        'Precio', 
        'Stock Total', 
        'Categoría',
        'Código Lote',
        'Cantidad Lote',
        'Fecha Caducidad',
        'Lote Activo'
    ]
    ws.append(encabezados)

    # Estilo para encabezados de columnas (AMARILLO)
    header_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, 10):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Datos de productos
    productos = Producto.objects.prefetch_related('lotes').all()
    row_num = 6

    # Color para filas de lote activo (AMARILLO CLARO)
    lote_activo_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')

    for producto in productos:
        lote_activo = producto.lotes.filter(
            cantidad__gt=0
        ).order_by('fecha_caducidad').first()
        
        lotes = producto.lotes.all()
        
        if lotes.exists():
            for lote in lotes:
                es_activo = "Sí" if lote_activo and lote.id == lote_activo.id else "No"
                ws.append([
                    producto.nombProduc,
                    producto.descripcion,
                    float(producto.precio),
                    producto.stock_total,
                    str(producto.Categoria),
                    lote.codigo_lote,
                    lote.cantidad,
                    lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'Sin fecha',
                    es_activo
                ])
                
                for col in range(1, 10):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = thin_border
                    if es_activo == "Sí":
                        cell.fill = lote_activo_fill
                
                row_num += 1
        else:
            ws.append([
                producto.nombProduc,
                producto.descripcion,
                float(producto.precio),
                producto.stock_total,
                str(producto.Categoria),
                'Sin lote',
                0,
                'N/A',
                'N/A'
            ])
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

    # Ajustar ancho de columnas
    column_widths = [25, 40, 12, 12, 20, 15, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Inventario_de_productos.xlsx'
    wb.save(response)
    return response


# ==================== CRUD CATEGORÍAS ====================

import logging

logger = logging.getLogger(__name__)

@admin_required
def agregar_categoria(request):
    try:
        logger.info("=" * 50)
        logger.info(f"🔹 REQUEST METHOD: {request.method}")
        logger.info(f"🔹 REQUEST PATH: {request.path}")
        logger.info(f"🔹 REQUEST ORIGIN: {request.META.get('HTTP_ORIGIN', 'No origin')}")
        logger.info(f"🔹 REQUEST REFERER: {request.META.get('HTTP_REFERER', 'No referer')}")
        logger.info(f"🔹 REMOTE_ADDR: {request.META.get('REMOTE_ADDR', 'No IP')}")
        logger.info(f"🔹 HTTP_X_FORWARDED_FOR: {request.META.get('HTTP_X_FORWARDED_FOR', 'No proxy IP')}")
        
        # CSRF Cookie - más seguro
        csrf_cookie = request.COOKIES.get('csrftoken', 'NO CSRF COOKIE')
        csrf_display = csrf_cookie[:20] + '...' if len(csrf_cookie) > 20 else csrf_cookie
        logger.info(f"🔹 CSRF Cookie: {csrf_display}")
        
        # Session Key - más seguro
        session_key = request.session.session_key
        session_display = session_key[:10] + '...' if session_key else 'NO SESSION'
        logger.info(f"🔹 Session Key: {session_display}")
        
        logger.info(f"🔹 User authenticated: {request.user.is_authenticated}")
        logger.info(f"🔹 User: {str(request.user) if request.user.is_authenticated else 'Anonymous'}")
        
        if request.method == "POST":
            logger.info(f"🔹 POST data keys: {list(request.POST.keys())}")
            logger.info(f"🔹 CSRF token in POST: {'csrfmiddlewaretoken' in request.POST}")
            
            nombre = request.POST.get("nombCategory")
            logger.info(f"🔹 Nombre de categoría: '{nombre}'")
            
            if nombre:
                try:
                    Category.objects.create(nombCategory=nombre)
                    logger.info(f"✅ Categoría '{nombre}' creada exitosamente")
                    messages.success(request, f"Categoría '{nombre}' agregada con éxito.")
                except Exception as e:
                    logger.error(f"❌ Error al crear categoría: {str(e)}")
                    messages.error(request, "Error al crear la categoría.")
            else:
                logger.warning("⚠️ Nombre de categoría vacío o None")
                messages.warning(request, "El nombre de la categoría no puede estar vacío.")
        else:
            logger.warning(f"⚠️ Método {request.method} no es POST")
        
        logger.info("=" * 50)
        
    except Exception as e:
        logger.error(f"💥 ERROR CRÍTICO en agregar_categoria: {str(e)}")
        logger.exception("Traceback completo:")
        messages.error(request, "Error interno del servidor.")
    
    return redirect('productos:list_product')

@admin_required
def editar_categoria(request, id):
    categoria = get_object_or_404(Category, id=id)
    if request.method == "POST":
        nuevo_nombre = request.POST.get("nombre")
        if nuevo_nombre:
            categoria.nombCategory = nuevo_nombre
            categoria.save()
    return redirect('productos:list_product')


@admin_required
def listar_categorias(request):
    categorias = Category.objects.all()
    productos = Producto.objects.all()
    form = CategoriaForm()

    if request.method == "POST":
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('productos:list_product')

    return render(request, 'productos/list_produc.html', {
        'categorias': categorias,
        'productos': productos,
        'form': form
    })


@admin_required
def eliminar_categoria(request, id):
    categoria = get_object_or_404(Category, id=id)
    categoria.delete()
    return redirect('productos:list_product')